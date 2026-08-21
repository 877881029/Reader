from pathlib import Path

from openpyxl import Workbook

from reader.formats.xlsx import to_html


def test_xlsx_renders_sheets_and_header(tmp_path: Path):
    path = tmp_path / "t.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Main"
    ws.append(["Name", "Value"])
    ws.append(["npu", "1"])
    ws2 = wb.create_sheet("Other")
    ws2.append(["X"])
    wb.save(path)
    result = to_html(path)
    assert "Main" in result.html
    assert "Other" in result.html
    assert "Name" in result.html
    assert "npu" in result.html
    assert 'id="sheet-Main"' in result.html


def test_xlsx_escapes_sheet_names_and_cell_values(tmp_path: Path):
    path = tmp_path / "escape.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = 'Sheet & "quotes"'
    ws.append(['<img onerror=alert(1)>'])
    wb.save(path)
    result = to_html(path)
    html = result.html
    assert "<img onerror=alert(1)>" not in html
    assert "&lt;img onerror=alert(1)&gt;" in html
    assert "Sheet &amp; &quot;quotes&quot;" in html


def test_xlsx_truncates_after_1000_rows(tmp_path: Path):
    path = tmp_path / "big.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Main"
    ws.append(["c"])
    for i in range(1005):
        ws.append([i])
    wb.save(path)
    result = to_html(path, max_rows=1000)
    assert "truncated" in result.html
    assert ">1004<" not in result.html
