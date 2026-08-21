from html import escape
from pathlib import Path

from openpyxl import load_workbook

from reader.preview.result import PreviewResult


def to_html(path: Path, max_rows: int = 1000) -> PreviewResult:
    wb = load_workbook(str(path), read_only=True, data_only=True)
    nav = ["<nav>"]
    bodies: list[str] = []
    for name in wb.sheetnames:
        nav.append(f'<a href="#sheet-{escape(name)}">{escape(name)}</a> ')
        ws = wb[name]
        rows_out = [f'<section id="sheet-{escape(name)}"><h2>{escape(name)}</h2><table>']
        truncated = False
        data_rows = 0
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            cells = [(c if c is not None else "") for c in row]
            tag = "th" if i == 0 else "td"
            if i > 0:
                data_rows += 1
                if data_rows > max_rows:
                    truncated = True
                    break
            tds = "".join(f"<{tag}>{escape(str(c))}</{tag}>" for c in cells)
            rows_out.append(f"<tr>{tds}</tr>")
        if truncated:
            rows_out.append("<p>truncated</p>")
        rows_out.append("</table></section>")
        bodies.append("".join(rows_out))
    nav.append("</nav>")
    wb.close()
    html = (
        "<!DOCTYPE html><html><head><meta charset='utf-8'></head><body>"
        + "".join(nav)
        + "".join(bodies)
        + "</body></html>"
    )
    return PreviewResult(html=html, status_label="内置预览", kind="html")
